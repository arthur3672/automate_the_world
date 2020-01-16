#!/usr/bin/python3

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import xml.etree.ElementTree as ET
import argparse
import numbers

def output_to_xlsx(output, output_file):
	workbook = load_workbook(output_file)
	ws = workbook.active
	ws.append(output)
	workbook.save(output_file)

def preparation(output_file):
	workbook = Workbook()
	ws = workbook.active
	ws.title = 'Items'

	ws.append(['No', 'Name', 'Description', 'CVSS Score', 'Severity', 'Proof of Concept', 'Affected Host', 'Remediation', 'Status'])
	bold_font = Font(bold=True)
	for cell in ws[1:1]:
		cell.font = bold_font
	workbook.save(filename=output_file)

def sort(input_file, output_file):
	input_workbook = load_workbook(input_file)
	input_ws = input_workbook.active
	preparation(output_file)
	score = 100
	number = 1
	while score > 0:
		cvss_score = score / 10
		for row in input_ws.rows:
			if row[3].value == cvss_score:
				output_workbook = load_workbook(output_file)
				output_ws = output_workbook.active
				row_value = [number, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value]
				output_ws.append(row_value)
				number += 1
				output_workbook.save(filename=output_file)
		score -= 1


def main(input_file, mid_file):
	tree = ET.parse(input_file)
	root = tree.getroot()[1]
	total_hosts = len(root)
	for i in range(0, total_hosts):
		target_ip = root[i].attrib['name']
		total_findings = len(root[i])
		for j in range(1, total_findings):
			duplicate = 0
			output = list()
			risk_factor = root[i][j].find('risk_factor').text
			if (risk_factor != 'None'):
				plugin_id = root[i][j].attrib['pluginID']
				workbook = load_workbook(filename=mid_file)
				ws = workbook['Items']
				row_count = ws.max_row + 1
				for row in range(2, row_count):
					if ws.cell(row=row, column=1).value == plugin_id:
						ws.cell(row=row, column=7).value += '\r\n' + str(target_ip)+':'+root[i][j].attrib['port']+' ('+root[i][j].attrib['protocol'].upper()+')'
						duplicate = 1
						workbook.save(filename=mid_file)
						break

				if duplicate == 0:				
					name = root[i][j].attrib['pluginName']
					description = root[i][j].find('description').text

					if root[i][j].find('cvss3_temporal_score') != None:
						cvss_score = root[i][j].find('cvss3_temporal_score').text
					elif root[i][j].find('cvss3_base_score') != None:
						cvss_score = root[i][j].find('cvss3_base_score').text
					elif root[i][j].find('cvss_temporal_score') != None:
						cvss_score = root[i][j].find('cvss_temporal_score').text
					elif root[i][j].find('cvss_base_score') != None:
						cvss_score = root[i][j].find('cvss_base_score').text
					else:
						if root[i][j].attrib['severity'] == 4:
							cvss_score = 10
						elif root[i][j].attrib['severity'] == 3:
							cvss_score = 8.9
						elif root[i][j].attrib['severity'] == 2:
							cvss_score = 6.9
						else:
							cvss_score = 3.9

					if float(cvss_score) < 4:
						severity = 'Low'
					elif float(cvss_score) < 7:
						severity = 'Medium'
					elif float(cvss_score) < 9:
						severity = 'High'
					else:
						severity = 'Critical'
					
					affected_host = str(target_ip)+':'+root[i][j].attrib['port']+' ('+root[i][j].attrib['protocol'].upper()+')'
					remediation = root[i][j].find('solution').text
					status = 'Not Solved'

					output = [plugin_id, name, description, float(cvss_score), severity, '', affected_host, remediation, status]
					output_to_xlsx(output, mid_file)
	
	output_file = mid_file[4:]
	sort(mid_file, output_file)
	print('Completed!')

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('-f', metavar='Nessus File / Directory', type=str, help='raw Nessus file / folder to generate XLSX report', required=True)
	parser.add_argument('-o', metavar='Output file', type=str, help='Output XLSX file', required=True)
	args = parser.parse_args()

	output_file = args.o
	mid_file = ''
	if '/' in output_file:
		mid_file = output_file.split('/')
		mid_file[-1] = 'mid_'+mid_file[-1]
		mid_file = '/'.join(mid_file)
	else:
		mid_file = 'mid_'+output_file
	preparation(mid_file)

	if os.path.isdir(args.f):
		for filename in os.listdir(args.f):
			if filename.endswith('.nessus'):
				fullpath_file = args.f + filename
				main(fullpath_file, mid_file)
	else:
		if args.f.endswith('.nessus'):
			main(args.f, mid_file)
		else:
			parser.print_usage()
	
	os.remove(mid_file)