#!/usr/bin/python3

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import xml.etree.ElementTree as ET
import argparse

def output_to_xlsx(target_os, output, output_file):
	workbook = load_workbook(output_file)
	ws = workbook[target_os]
	current_row_count = ws.max_row
	output.insert(0, current_row_count)
	ws.append(output)
	workbook.save(filename=output_file)

def preparation(target_os, output_file):
	workbook = load_workbook(output_file)
	workbook.create_sheet(target_os)
	ws = workbook[target_os]
	ws.append(['No', 'Name', 'Description', 'Severity', 'Current Host Value', 'Affected Host', 'Remediation', 'Status'])
	bold_font = Font(bold=True)
	for cell in ws[1:1]:
		cell.font = bold_font
	workbook.save(filename=output_file)

def main(input_file, output_file):
	
	tree = ET.parse(input_file)
	root = tree.getroot()[1]
	total_hosts = len(root)
	output = list()
	for i in range(0, total_hosts):
		target_ip = root[i].attrib['name']
		total_findings = len(root[i])
		if total_findings > 0:
			j = total_findings - 1
			while j > 0:
				if int(root[i][j].attrib['pluginID']) == 21156:
					severity = root[i][j].find('{http://www.nessus.org/cm}compliance-result').text.capitalize()
					if severity == 'Failed':
						os_type = root[i][j].find('{http://www.nessus.org/cm}compliance-audit-file').text.split('_')[1:]
						target_os = ''
						for os in os_type:
							try:
								int(os)
								target_os += os
								break
							except:
								target_os += os + ' '

						workbook = load_workbook(output_file)
						if target_os not in workbook.sheetnames:
							preparation(target_os, output_file)
						
						element_name = str(list(root[i][j]))
						workbook = load_workbook(output_file)
						ws = workbook[target_os]
						duplicate = 0
						for x in range(2, ws.max_row):
							if ws.cell(row=x, column=2).value == root[i][j].find('{http://www.nessus.org/cm}compliance-check-name').text:
								ws.cell(row=x, column=5).value += '\r\nFor ' + str(target_ip) + ':\r\n' + root[i][j].find('{http://www.nessus.org/cm}compliance-actual-value').text
								ws.cell(row=x, column=6).value += '\r\n'+target_ip
								workbook.save(output_file)
								duplicate = 1
								break
						if duplicate == 0:
							if 'compliance-check-name' in element_name:
								name = root[i][j].find('{http://www.nessus.org/cm}compliance-check-name').text
								name = name.split(' ')[1:]
								name = ' '.join(name)
							else:
								name = ''

							if 'compliance-info' in element_name:
								description = root[i][j].find('{http://www.nessus.org/cm}compliance-info').text
								description = description[1:-1]
							else:
								description = ''

							current_value = 'For ' + str(target_ip) + ':\r\n'
							if 'compliance-actual-value' in element_name:
								current_value += root[i][j].find('{http://www.nessus.org/cm}compliance-actual-value').text
							else:
								current_value += 'No Value for this IP'

							if 'compliance-solution' in element_name:
								solution = root[i][j].find('{http://www.nessus.org/cm}compliance-solution').text
							else:
								solution = ''

							status = 'Not Solved'
							output = [name, description, severity, current_value, target_ip, solution, status]
							output_to_xlsx(target_os, output, output_file)
				j -= 1
	print('Completed!')

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('-f', metavar='Nessus File / Directory', type=str, help='raw Nessus file / folder to generate XLSX report', required=True)
	parser.add_argument('-o', metavar='Output file', type=str, help='Output XLSX file', required=True)
	args = parser.parse_args()

	if args.o.endswith('.xlsx'):
		workbook = Workbook()
		workbook.save(filename=args.o)
		if os.path.isdir(args.f):
			for filename in os.listdir(args.f):
				if filename.endswith('.nessus'):
					fullpath_file = args.f + filename
					main(fullpath_file, args.o)
		else:
			main(args.f, args.o)
		
		workbook = load_workbook(args.o)
		workbook.remove('Sheet')
		workbook.save(filename=args.o)
	else:
		parser.print_usage()
	
	exit()

