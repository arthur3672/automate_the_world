#!/usr/bin/python3

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import xml.etree.ElementTree as ET
import argparse

def output_to_xlsx(target_ip, output, output_file):
	workbook = load_workbook(output_file)
	ws = workbook[target_ip]
	current_row_count = ws.max_row
	output.insert(0, current_row_count)
	ws.append(output)
	workbook.save(filename=output_file)

def preparation(target_ip, output_file):
	workbook = load_workbook(output_file)
	workbook.create_sheet(target_ip)
	ws = workbook[target_ip]
	ws.append(['No', 'Name', 'Description', 'Severity', 'Current Host Value', 'Remediation', 'Status'])
	bold_font = Font(bold=True)
	for cell in ws[1:1]:
		cell.font = bold_font
	workbook.save(filename=output_file)

def main(input_file, output_file):
	tree = ET.parse(input_file)
	root = tree.getroot()[1]
	total_hosts = len(root)
	output = list()
	for i in range(0, total_hosts):
		target_ip = root[i].attrib['name']
		total_findings = len(root[i])
		if total_findings > 0:
			preparation(target_ip, output_file)
			j = total_findings - 1
			while j > 0:
				if int(root[i][j].attrib['pluginID']) == 33814:
					severity = root[i][j].find('{http://www.nessus.org/cm}compliance-result').text.capitalize()
					if severity == 'Failed':
						element_name = str(list(root[i][j]))
						if 'compliance-check-name' in element_name:
							name = root[i][j].find('{http://www.nessus.org/cm}compliance-check-name').text
							name = name.split(' ')[1:]
							name = ' '.join(name)
						else:
							name = 'Findings name not available'

						if 'compliance-info' in element_name:
							description = root[i][j].find('{http://www.nessus.org/cm}compliance-info').text
						else:
							description = 'Description not available'

						if 'compliance-actual-value' in element_name:
							current_value = root[i][j].find('{http://www.nessus.org/cm}compliance-actual-value').text
						else:
							current_value = 'No Value for this IP'

						if 'compliance-solution' in element_name:
							solution = root[i][j].find('{http://www.nessus.org/cm}compliance-solution').text
						else:
							solution = 'Solution not available'

						status = 'Not Solved'
						output = [name, description, severity, current_value, solution, status]
						output_to_xlsx(target_ip, output, output_file)
				j -= 1

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('-f', metavar='Nessus File / Directory', type=str, help='raw Nessus file / folder to generate XLSX report', required=True)
	parser.add_argument('-o', metavar='Output file', type=str, help='Output XLSX file', required=True)
	args = parser.parse_args()

	if args.o.endswith('.xlsx'):
		workbook = Workbook()
		workbook.save(filename=args.o)
		if os.path.isdir(args.f):
			for filename in os.listdir(args.f):
				if filename.endswith('.nessus'):
					fullpath_file = args.f + filename
					main(fullpath_file, args.o)
		else:
			main(args.f, args.o)
		workbook = load_workbook(filename=args.o)
		workbook.remove(workbook['Sheet'])
		workbook.save(filename=args.o)
		print('Completed!')
	else:
		parser.print_usage()
